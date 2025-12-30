import csv
import io
import json
import os
from decimal import Decimal
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from apps.accounts.decorators import login_required
from apps.core.htmx import htmx_view
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, F
from django.db import models, transaction, IntegrityError
from django.conf import settings
from apps.configuration.models import HubConfig, StoreConfig, TaxClass
from .models import Product, Category


@login_required
@htmx_view('inventory/pages/index.html', 'inventory/partials/dashboard_content.html')
def dashboard(request):
    """Dashboard principal del inventario con estadísticas y resumen"""
    # Obtener configuración global
    currency = HubConfig.get_value('currency', 'EUR')

    # Estadísticas
    total_products = Product.objects.filter(is_active=True).count()
    products_in_stock = Product.objects.filter(is_active=True, stock__gt=0).count()
    products_low_stock = Product.objects.filter(is_active=True).filter(
        stock__lte=F('low_stock_threshold')
    ).count()
    total_inventory_value = Product.objects.filter(is_active=True).aggregate(
        total=Sum(F('stock') * F('price'))
    )['total'] or 0

    # Categorías (top 5 con conteo de productos)
    categories = Category.objects.filter(is_active=True).order_by('order')[:5]

    # Productos con stock bajo (top 10)
    low_stock_products = Product.objects.filter(
        is_active=True,
        stock__lte=F('low_stock_threshold')
    ).order_by('stock')[:10]

    return {
        'current_view': 'dashboard',
        'current_section': 'inventory',
        'total_products': total_products,
        'products_in_stock': products_in_stock,
        'products_low_stock': products_low_stock,
        'total_inventory_value': total_inventory_value,
        'currency': currency,
        'categories': categories,
        'low_stock_products': low_stock_products,
    }


@login_required
def products_list(request):
    """Lista de productos con infinite scroll"""
    from .models import ProductsConfig
    from apps.core.htmx import InfiniteScrollPaginator

    # Filtrar productos
    queryset = Product.objects.filter(is_active=True).prefetch_related('categories')

    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(name__icontains=search) |
            Q(sku__icontains=search) |
            Q(categories__name__icontains=search)
        ).distinct()

    # Ordenamiento
    allowed_order_fields = ['name', '-name', 'sku', '-sku', 'price', '-price', 'stock', '-stock', 'id', '-id']
    order_by = request.GET.get('order_by', '-id')
    if order_by in allowed_order_fields:
        queryset = queryset.order_by(order_by)
    else:
        queryset = queryset.order_by('-id')

    # Infinite scroll pagination
    per_page = int(request.GET.get('per_page', '25'))
    paginator = InfiniteScrollPaginator(queryset, per_page=per_page)
    page_data = paginator.get_page(request.GET.get('page', 1))

    config = ProductsConfig.get_config()

    context = {
        'current_view': 'products',
        'current_section': 'inventory',
        'products': page_data['items'],
        'has_next': page_data['has_next'],
        'next_page': page_data['next_page'],
        'is_first_page': page_data['is_first_page'],
        'total_count': page_data['total_count'],
        'start_index': page_data['start_index'],
        'end_index': page_data['end_index'],
        'page_number': page_data['page_number'],
        'barcode_enabled': config.barcode_enabled,
        'search': search,
        'order_by': order_by,
        'per_page': per_page,
    }

    # Determine which template to render
    is_htmx = request.headers.get('HX-Request') or request.GET.get('partial') == 'true'
    page_num = page_data['page_number']
    hx_target = request.headers.get('HX-Target', '')

    if not is_htmx:
        return render(request, 'inventory/pages/products.html', context)

    if page_num > 1:
        # Subsequent pages - only return table rows + loader
        return render(request, 'inventory/partials/products_rows_infinite.html', context)

    # Check if targeting table container (search/sort) vs full content (tab navigation)
    if hx_target == 'products-table-container':
        # Search/sort - only return table partial
        return render(request, 'inventory/partials/products_table_partial.html', context)

    # Tab navigation - return full content partial
    return render(request, 'inventory/partials/products_content.html', context)


@login_required
def product_list_ajax(request):
    """
    Retorna la lista de productos en formato JSON para tabla dinámica
    """
    search = request.GET.get('search', '')
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))

    # Filtros
    products_queryset = Product.objects.filter(is_active=True)

    if search:
        products_queryset = products_queryset.filter(
            Q(name__icontains=search) |
            Q(sku__icontains=search) |
            Q(categories__name__icontains=search)
        ).distinct()

    # Paginación
    paginator = Paginator(products_queryset, per_page)
    page_obj = paginator.get_page(page)

    products_data = []
    for product in page_obj:
        categories_list = [{'id': cat.id, 'name': cat.name} for cat in product.categories.all()]
        products_data.append({
            'id': product.id,
            'name': product.name,
            'sku': product.sku,
            'categories': categories_list,
            'category': categories_list[0]['name'] if categories_list else 'Sin categoría',
            'category_ids': [cat['id'] for cat in categories_list],
            'price': float(product.price),
            'cost': float(product.cost),
            'stock': product.stock,
            'low_stock_threshold': product.low_stock_threshold,
            'is_low_stock': product.is_low_stock,
            'image': product.get_image_path(),
            'initial': product.get_initial(),
            'profit_margin': float(product.profit_margin),
        })

    return JsonResponse({
        'products': products_data,
        'total': paginator.count,
        'pages': paginator.num_pages,
        'current_page': page_obj.number,
    })


@login_required
@require_http_methods(["GET", "POST"])
def product_create(request):
    """
    Crear un nuevo producto
    """
    if request.method == 'POST':
        try:
            # Procesar imagen
            image = request.FILES.get('image')

            # Validar y convertir price
            price_str = request.POST.get('price', '').strip()
            if not price_str:
                raise ValueError("El precio es requerido")
            price = Decimal(price_str)

            # Validar y convertir cost
            cost_str = request.POST.get('cost', '0').strip()
            cost = Decimal(cost_str) if cost_str else Decimal('0')

            # Validar y convertir stock
            stock_str = request.POST.get('stock', '0').strip()
            stock = int(stock_str) if stock_str else 0

            # Validar y convertir low_stock_threshold
            threshold_str = request.POST.get('low_stock_threshold', '10').strip()
            low_stock_threshold = int(threshold_str) if threshold_str else 10

            # Get product_type and tax_class
            product_type = request.POST.get('product_type', 'physical')
            tax_class_id = request.POST.get('tax_class_id', '').strip()
            tax_class = TaxClass.objects.get(id=tax_class_id) if tax_class_id else None

            # Create product first
            product = Product.objects.create(
                name=request.POST['name'],
                sku=request.POST['sku'],
                description=request.POST.get('description', ''),
                product_type=product_type,
                price=price,
                cost=cost,
                stock=stock,
                low_stock_threshold=low_stock_threshold,
                tax_class=tax_class,
                image=image
            )

            # Handle multiple categories by name (support both formats)
            category_names = request.POST.getlist('category_names[]') or request.POST.getlist('category_names')
            # If we get a single string with comma-separated values, split it
            if len(category_names) == 1 and ',' in category_names[0]:
                category_names = category_names[0].split(',')
            # Filter out empty strings and capitalize
            category_names = [name.strip().capitalize() for name in category_names if name.strip()]
            if category_names:
                categories = Category.objects.filter(name__in=category_names)
                product.categories.set(categories)
            # If no categories provided, leave empty (product can have no category)

            # Redirect to products list
            return redirect('inventory:products_list')
        except IntegrityError as e:
            # Handle duplicate SKU error
            error_msg = str(e)
            if 'sku' in error_msg.lower() or 'unique' in error_msg.lower():
                error_message = "Ya existe un producto con ese SKU. Por favor, use un SKU diferente."
            else:
                error_message = "Error de integridad en la base de datos."
            categories = Category.objects.filter(is_active=True).order_by('order', 'name')
            tax_classes = TaxClass.objects.filter(is_active=True).order_by('order', 'rate')
            context = {
                'categories': categories,
                'tax_classes': tax_classes,
                'mode': 'create',
                'readonly': False,
                'error_message': error_message
            }
            return render(request, 'inventory/product_form.html', context)
        except Exception as e:
            # Show error message and return to form
            categories = Category.objects.filter(is_active=True).order_by('order', 'name')
            tax_classes = TaxClass.objects.filter(is_active=True).order_by('order', 'rate')
            context = {
                'categories': categories,
                'tax_classes': tax_classes,
                'mode': 'create',
                'readonly': False,
                'error_message': str(e)
            }
            return render(request, 'inventory/product_form.html', context)

    # GET request - show create form
    categories = Category.objects.filter(is_active=True).order_by('order', 'name')
    tax_classes = TaxClass.objects.filter(is_active=True).order_by('order', 'rate')
    context = {
        'categories': categories,
        'tax_classes': tax_classes,
        'mode': 'create',
        'readonly': False,
    }
    return render(request, 'inventory/product_form.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def product_edit(request, pk):
    """
    Editar un producto existente
    """
    product = get_object_or_404(Product, pk=pk)

    if request.method == 'POST':
        try:
            # Validar y convertir price
            price_str = request.POST.get('price', '').strip()
            if not price_str:
                raise ValueError("El precio es requerido")

            # Validar y convertir cost
            cost_str = request.POST.get('cost', '0').strip()

            # Validar y convertir stock
            stock_str = request.POST.get('stock', '0').strip()

            # Validar y convertir low_stock_threshold
            threshold_str = request.POST.get('low_stock_threshold', '10').strip()

            product.name = request.POST['name']
            product.sku = request.POST['sku']
            product.description = request.POST.get('description', '')
            product.product_type = request.POST.get('product_type', 'physical')
            product.price = Decimal(price_str)
            product.cost = Decimal(cost_str) if cost_str else Decimal('0')
            product.stock = int(stock_str) if stock_str else 0
            product.low_stock_threshold = int(threshold_str) if threshold_str else 10

            # Update tax_class
            tax_class_id = request.POST.get('tax_class_id', '').strip()
            product.tax_class = TaxClass.objects.get(id=tax_class_id) if tax_class_id else None

            # Actualizar imagen si se proporciona
            if 'image' in request.FILES:
                # Eliminar imagen anterior
                if product.image:
                    if os.path.isfile(product.image.path):
                        os.remove(product.image.path)
                product.image = request.FILES['image']

            product.save()

            # Update multiple categories by name (support both formats)
            category_names = request.POST.getlist('category_names[]') or request.POST.getlist('category_names')
            # If we get a single string with comma-separated values, split it
            if len(category_names) == 1 and ',' in category_names[0]:
                category_names = category_names[0].split(',')
            # Filter out empty strings and capitalize
            category_names = [name.strip().capitalize() for name in category_names if name.strip()]
            if category_names:
                categories = Category.objects.filter(name__in=category_names)
                product.categories.set(categories)
            else:
                product.categories.clear()  # Remove all categories if none selected

            # Redirect to products list
            return redirect('inventory:products_list')
        except IntegrityError as e:
            # Handle duplicate SKU error
            error_msg = str(e)
            if 'sku' in error_msg.lower() or 'unique' in error_msg.lower():
                error_message = "Ya existe un producto con ese SKU. Por favor, use un SKU diferente."
            else:
                error_message = "Error de integridad en la base de datos."
            categories = Category.objects.filter(is_active=True).order_by('order', 'name')
            tax_classes = TaxClass.objects.filter(is_active=True).order_by('order', 'rate')
            context = {
                'product': product,
                'categories': categories,
                'tax_classes': tax_classes,
                'mode': 'edit',
                'readonly': False,
                'error_message': error_message
            }
            return render(request, 'inventory/product_form.html', context)
        except Exception as e:
            # Show error message and return to form
            categories = Category.objects.filter(is_active=True).order_by('order', 'name')
            tax_classes = TaxClass.objects.filter(is_active=True).order_by('order', 'rate')
            context = {
                'product': product,
                'categories': categories,
                'tax_classes': tax_classes,
                'mode': 'edit',
                'readonly': False,
                'error_message': str(e)
            }
            return render(request, 'inventory/product_form.html', context)

    # GET request - show edit form
    categories = Category.objects.filter(is_active=True).order_by('order', 'name')
    tax_classes = TaxClass.objects.filter(is_active=True).order_by('order', 'rate')
    context = {
        'product': product,
        'categories': categories,
        'tax_classes': tax_classes,
        'mode': 'edit',
        'readonly': False,
    }
    return render(request, 'inventory/product_form.html', context)


@login_required
@require_http_methods(["GET"])
def product_view(request, pk):
    """
    Ver un producto en modo solo lectura
    """
    product = get_object_or_404(Product, pk=pk)
    categories = Category.objects.filter(is_active=True).order_by('order', 'name')
    tax_classes = TaxClass.objects.filter(is_active=True).order_by('order', 'rate')

    context = {
        'product': product,
        'categories': categories,
        'tax_classes': tax_classes,
        'mode': 'view',
        'readonly': True,
    }
    return render(request, 'inventory/product_form.html', context)


@login_required
@require_http_methods(["POST", "DELETE"])
def product_delete(request, pk):
    """
    Eliminar un producto
    """
    try:
        product = get_object_or_404(Product, pk=pk)
        product.delete()

        return JsonResponse({
            'success': True,
            'message': 'Producto eliminado exitosamente'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
def export_csv(request):
    """
    Exportar productos a CSV con nombres de campo del modelo (inglés)
    """
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="productos.csv"'
    response.write('\ufeff')  # BOM para UTF-8

    writer = csv.writer(response)
    # Usar nombres de campo del modelo (field names en inglés)
    writer.writerow(['sku', 'name', 'description', 'categories', 'price', 'cost', 'stock', 'low_stock_threshold', 'ean13'])

    products = Product.objects.filter(is_active=True)
    for product in products:
        categories_names = ', '.join([cat.name for cat in product.categories.all()]) if product.categories.exists() else ''
        writer.writerow([
            product.sku,
            product.name,
            product.description,
            categories_names,
            float(product.price),
            float(product.cost),
            product.stock,
            product.low_stock_threshold,
            product.ean13 or '',
        ])

    return response


@login_required
@require_http_methods(["POST"])
def import_csv(request):
    """
    Importar productos desde CSV
    """
    try:
        csv_file = request.FILES.get('file')

        if not csv_file:
            return JsonResponse({
                'success': False,
                'message': 'No se proporcionó ningún archivo'
            }, status=400)

        if not csv_file.name.endswith('.csv'):
            return JsonResponse({
                'success': False,
                'message': 'El archivo debe ser CSV'
            }, status=400)

        # Leer CSV
        decoded_file = csv_file.read().decode('utf-8-sig')
        io_string = io.StringIO(decoded_file)
        reader = csv.DictReader(io_string)

        created = 0
        updated = 0
        errors = []

        with transaction.atomic():
            for row in reader:
                try:
                    # Usar nombres de campo exactos de la base de datos (inglés)
                    sku = row.get('sku', '').strip()
                    if not sku:
                        continue

                    # Validations
                    name = row.get('name', '').strip()
                    if not name:
                        errors.append(f"Fila {reader.line_num}: El nombre es obligatorio")
                        continue

                    price = Decimal(row.get('price', '0'))
                    cost = Decimal(row.get('cost', '0'))
                    stock = int(row.get('stock', '0'))

                    if price < 0:
                        errors.append(f"Fila {reader.line_num}: El precio no puede ser negativo")
                        continue

                    if cost < 0:
                        errors.append(f"Fila {reader.line_num}: El costo no puede ser negativo")
                        continue

                    if stock < 0:
                        errors.append(f"Fila {reader.line_num}: El stock no puede ser negativo")
                        continue

                    # Get or create categories (comma-separated)
                    categories_str = row.get('categories', '').strip()
                    categories_list = []
                    if categories_str:
                        category_names = [cat_name.strip() for cat_name in categories_str.split(',')]
                        for cat_name in category_names:
                            if cat_name:
                                category, _ = Category.objects.get_or_create(
                                    name=cat_name,
                                    defaults={'icon': 'cube-outline', 'is_active': True}
                                )
                                categories_list.append(category)

                    product, is_created = Product.objects.update_or_create(
                        sku=sku,
                        defaults={
                            'name': name,
                            'description': row.get('description', ''),
                            'price': price,
                            'cost': cost,
                            'stock': stock,
                            'low_stock_threshold': int(row.get('low_stock_threshold', '10')),
                            'ean13': row.get('ean13', '') or None,
                        }
                    )

                    # Set categories
                    if categories_list:
                        product.categories.set(categories_list)

                    if is_created:
                        created += 1
                    else:
                        updated += 1

                except Exception as e:
                    errors.append(f"Error en fila {reader.line_num}: {str(e)}")

        return JsonResponse({
            'success': True,
            'message': f'{created} productos creados, {updated} actualizados',
            'created': created,
            'updated': updated,
            'errors': errors
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al importar: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["POST"])
def import_excel(request):
    """
    Importar productos desde Excel
    Requiere openpyxl: pip install openpyxl
    """
    try:
        excel_file = request.FILES.get('file')

        if not excel_file:
            return JsonResponse({
                'success': False,
                'message': 'No se proporcionó ningún archivo'
            }, status=400)

        if not (excel_file.name.endswith('.xlsx') or excel_file.name.endswith('.xls')):
            return JsonResponse({
                'success': False,
                'message': 'El archivo debe ser Excel (.xlsx o .xls)'
            }, status=400)

        try:
            import openpyxl
        except ImportError:
            return JsonResponse({
                'success': False,
                'message': 'Librería openpyxl no instalada. Ejecute: pip install openpyxl'
            }, status=400)

        # Leer Excel
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # Obtener encabezados (primera fila)
        headers = [cell.value for cell in ws[1]]

        created = 0
        updated = 0
        errors = []

        with transaction.atomic():
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    row_dict = dict(zip(headers, row))

                    sku = str(row_dict.get('SKU', '')).strip()
                    if not sku:
                        continue

                    # Validations
                    name = str(row_dict.get('Nombre', '')).strip()
                    if not name:
                        errors.append(f"Fila {row_idx}: El nombre es obligatorio")
                        continue

                    price = Decimal(str(row_dict.get('Precio', '0')))
                    cost = Decimal(str(row_dict.get('Costo', '0')))
                    stock = int(row_dict.get('Stock', 0))

                    if price < 0:
                        errors.append(f"Fila {row_idx}: El precio no puede ser negativo")
                        continue

                    if cost < 0:
                        errors.append(f"Fila {row_idx}: El costo no puede ser negativo")
                        continue

                    if stock < 0:
                        errors.append(f"Fila {row_idx}: El stock no puede ser negativo")
                        continue

                    # Get or create categories (comma-separated)
                    categories_str = str(row_dict.get('Categoría', '')).strip()
                    categories_list = []
                    if categories_str:
                        category_names = [name.strip() for name in categories_str.split(',')]
                        for cat_name in category_names:
                            if cat_name:
                                category, _ = Category.objects.get_or_create(
                                    name=cat_name,
                                    defaults={'icon': 'cube-outline', 'is_active': True}
                                )
                                categories_list.append(category)

                    product, is_created = Product.objects.update_or_create(
                        sku=sku,
                        defaults={
                            'name': name,
                            'description': str(row_dict.get('Descripción', '')),
                            'price': price,
                            'cost': cost,
                            'stock': stock,
                            'low_stock_threshold': int(row_dict.get('Umbral Stock Bajo', 10)),
                        }
                    )

                    # Set categories
                    if categories_list:
                        product.categories.set(categories_list)

                    if is_created:
                        created += 1
                    else:
                        updated += 1

                except Exception as e:
                    errors.append(f"Error en fila {row_idx}: {str(e)}")

        return JsonResponse({
            'success': True,
            'message': f'{created} productos creados, {updated} actualizados',
            'created': created,
            'updated': updated,
            'errors': errors
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al importar: {str(e)}'
        }, status=400)


@login_required
def categories_list(request):
    """
    Retorna la lista de categorías con búsqueda y paginación.
    Parámetros opcionales:
    - search: texto a buscar en nombre
    - id: filtrar por ID específico
    - page: número de página (default: 1)
    - per_page: resultados por página (default: 10)
    """
    from django.db.models import Q

    # Base queryset
    categories = Category.objects.filter(is_active=True)

    # Filtro por ID (para editar categoría)
    category_id = request.GET.get('id')
    if category_id:
        categories = categories.filter(id=category_id)

    # Búsqueda
    search = request.GET.get('search', '').strip()
    if search:
        categories = categories.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )

    # Orden
    categories = categories.order_by('order', 'name')

    # Total antes de paginar
    total = categories.count()

    # Paginación
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))

    # Calcular offset
    offset = (page - 1) * per_page
    categories = categories[offset:offset + per_page]

    # Serializar datos
    categories_data = [
        {
            'id': category.id,
            'name': category.name,
            'icon': category.icon,
            'color': category.color,
            'description': category.description,
            'order': category.order,
            'image': category.get_image_url(),
            'initial': category.get_initial(),
            'product_count': category.product_count,
        }
        for category in categories
    ]

    return JsonResponse({
        'success': True,
        'categories': categories_data,
        'total': total
    })


@login_required
def categories_index(request):
    """Vista principal de gestión de categorías con infinite scroll."""
    from apps.core.htmx import InfiniteScrollPaginator

    # Filtrar categorías
    queryset = Category.objects.filter(is_active=True)

    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )

    # Ordenamiento
    order_by = request.GET.get('order_by', 'order')
    queryset = queryset.order_by(order_by, 'name')

    # Infinite scroll pagination
    per_page = int(request.GET.get('per_page', 25))
    paginator = InfiniteScrollPaginator(queryset, per_page=per_page)
    page_data = paginator.get_page(request.GET.get('page', 1))

    context = {
        'current_view': 'categories',
        'current_section': 'inventory',
        'categories': page_data['items'],
        'has_next': page_data['has_next'],
        'next_page': page_data['next_page'],
        'is_first_page': page_data['is_first_page'],
        'total_count': page_data['total_count'],
        'search': search,
        'order_by': order_by,
        'per_page': per_page,
    }

    # Determine which template to render
    is_htmx = request.headers.get('HX-Request') or request.GET.get('partial') == 'true'
    page_num = page_data['page_number']
    hx_target = request.headers.get('HX-Target', '')

    if not is_htmx:
        return render(request, 'inventory/pages/categories.html', context)

    if page_num > 1:
        # Subsequent pages - only return table rows + loader
        return render(request, 'inventory/partials/categories_rows_infinite.html', context)

    # Check if targeting table container (search/sort) vs full content (tab navigation)
    if hx_target == 'categories-table-container':
        # Search/sort - only return table partial
        return render(request, 'inventory/partials/categories_table_partial.html', context)

    # Tab navigation - return full content partial
    return render(request, 'inventory/partials/categories_content.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def category_create(request):
    """Crear una nueva categoría."""
    if request.method == 'POST':
        try:
            image = request.FILES.get('image')

            # Handle tax_class
            tax_class = None
            tax_class_id = request.POST.get('tax_class_id', '').strip()
            if tax_class_id:
                tax_class = TaxClass.objects.filter(id=tax_class_id, is_active=True).first()

            category = Category.objects.create(
                name=request.POST['name'],
                description=request.POST.get('description', ''),
                icon=request.POST.get('icon', 'cube-outline'),
                color=request.POST.get('color', '#3880ff'),
                order=int(request.POST.get('order', 0)),
                image=image,
                tax_class=tax_class
            )

            return JsonResponse({
                'success': True,
                'message': 'Categoría creada exitosamente',
                'category_id': category.id
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)

    # Get tax classes for the form
    tax_classes = TaxClass.objects.filter(is_active=True).order_by('order', 'name')
    return render(request, 'inventory/category_form.html', {
        'tax_classes': tax_classes,
        'mode': 'create'
    })


@login_required
@require_http_methods(["GET", "POST"])
def category_edit(request, pk):
    """Editar una categoría existente."""
    category = get_object_or_404(Category, pk=pk)

    if request.method == 'POST':
        try:
            category.name = request.POST['name']
            category.description = request.POST.get('description', '')
            category.icon = request.POST.get('icon', 'cube-outline')
            category.color = request.POST.get('color', '#3880ff')
            category.order = int(request.POST.get('order', 0))

            # Handle tax_class
            tax_class_id = request.POST.get('tax_class_id', '').strip()
            if tax_class_id:
                category.tax_class = TaxClass.objects.filter(id=tax_class_id, is_active=True).first()
            else:
                category.tax_class = None

            # Actualizar imagen si se proporciona
            if 'image' in request.FILES:
                if category.image:
                    if os.path.isfile(category.image.path):
                        os.remove(category.image.path)
                category.image = request.FILES['image']

            category.save()

            return JsonResponse({
                'success': True,
                'message': 'Categoría actualizada exitosamente'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)

    # Get tax classes for the form
    tax_classes = TaxClass.objects.filter(is_active=True).order_by('order', 'name')
    context = {
        'category': category,
        'tax_classes': tax_classes,
        'mode': 'edit'
    }
    return render(request, 'inventory/category_form.html', context)


@login_required
@require_http_methods(["POST"])
def category_delete(request, pk):
    """Eliminar una categoría."""
    try:
        category = get_object_or_404(Category, pk=pk)

        # Check if category has products
        product_count = category.products.filter(is_active=True).count()
        if product_count > 0:
            return JsonResponse({
                'success': False,
                'message': f'No se puede eliminar la categoría porque tiene {product_count} productos asociados'
            }, status=400)

        category.delete()

        return JsonResponse({
            'success': True,
            'message': 'Categoría eliminada exitosamente'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
@htmx_view('inventory/pages/reports.html', 'inventory/partials/reports_content.html')
def reports_view(request):
    """Vista de informes/reportes del inventario"""
    # Estadísticas generales
    total_products = Product.objects.filter(is_active=True).count()
    products_in_stock = Product.objects.filter(is_active=True, stock__gt=0).count()
    products_out_of_stock = Product.objects.filter(is_active=True, stock=0).count()
    products_low_stock = Product.objects.filter(
        is_active=True,
        stock__lte=F('low_stock_threshold'),
        stock__gt=0
    ).count()

    # Valor total del inventario
    total_inventory_value = Product.objects.filter(is_active=True).aggregate(
        total=Sum(F('stock') * F('price'))
    )['total'] or 0

    # Valor total del costo
    total_cost_value = Product.objects.filter(is_active=True).aggregate(
        total=Sum(F('stock') * F('cost'))
    )['total'] or 0

    # Unidades totales
    total_units = Product.objects.filter(is_active=True).aggregate(
        total=Sum('stock')
    )['total'] or 0

    # Total de categorías
    total_categories = Category.objects.filter(is_active=True).count()

    # Estadísticas por categoría
    category_stats = []
    categories = Category.objects.filter(is_active=True).order_by('order', 'name')
    for category in categories:
        products = Product.objects.filter(is_active=True, categories=category)
        product_count = products.count()
        if product_count > 0:
            total_stock = products.aggregate(total=Sum('stock'))['total'] or 0
            total_value = products.aggregate(
                total=Sum(F('stock') * F('price'))
            )['total'] or 0
            category_stats.append({
                'name': category.name,
                'icon': category.icon,
                'color': category.color,
                'product_count': product_count,
                'total_stock': total_stock,
                'total_value': total_value,
            })

    # Top productos por valor en stock
    top_products_by_value = Product.objects.filter(
        is_active=True,
        stock__gt=0
    ).annotate(
        stock_value=F('stock') * F('price')
    ).order_by('-stock_value')[:10]

    # Top productos por cantidad en stock
    top_products_by_stock = Product.objects.filter(
        is_active=True,
        stock__gt=0
    ).order_by('-stock')[:10]

    # Productos en stock crítico (stock menor o igual al umbral)
    critical_stock_products = Product.objects.filter(
        is_active=True,
        stock__lte=F('low_stock_threshold'),
        stock__gt=0
    ).order_by('stock')[:20]

    return {
        'current_view': 'reports',
        'current_section': 'inventory',
        'total_products': total_products,
        'products_in_stock': products_in_stock,
        'products_out_of_stock': products_out_of_stock,
        'products_low_stock': products_low_stock,
        'total_inventory_value': total_inventory_value,
        'total_cost_value': total_cost_value,
        'total_units': total_units,
        'total_categories': total_categories,
        'category_stats': category_stats,
        'top_products_by_value': top_products_by_value,
        'top_products_by_stock': top_products_by_stock,
        'critical_stock_products': critical_stock_products,
    }


@login_required
def settings_view(request):
    """Vista de configuración del plugin de inventario"""
    import json
    from .models import ProductsConfig

    config = ProductsConfig.get_config()

    if request.method == "POST":
        try:
            # Parse JSON body
            data = json.loads(request.body)

            # Update config
            config.allow_negative_stock = data.get('allow_negative_stock', False)
            config.low_stock_alert_enabled = data.get('low_stock_alert_enabled', True)
            config.barcode_enabled = data.get('barcode_enabled', True)
            config.save()

            return JsonResponse({"success": True, "message": "Configuración guardada correctamente"})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    context = {
        'current_view': 'settings',
        'current_section': 'inventory',
        "config": config,
    }

    # Use htmx pattern manually since POST needs special handling
    if request.headers.get('HX-Request'):
        return render(request, "inventory/partials/settings_content.html", context)
    return render(request, "inventory/pages/settings.html", context)


@login_required
@require_http_methods(["GET"])
def generate_barcode(request, product_id):
    """
    Generate barcode SVG for a product
    Returns SVG image directly
    Supports: ?type=sku (CODE128) or ?type=ean13 (EAN-13)
    """
    from .barcode_utils import generate_barcode_svg
    from .models import ProductsConfig

    # Check if barcode generation is enabled
    config = ProductsConfig.get_config()
    if not config.barcode_enabled:
        return JsonResponse({
            "success": False,
            "error": "Barcode generation is disabled in settings"
        }, status=403)

    try:
        product = get_object_or_404(Product, id=product_id)

        # Get barcode type from query parameter (default to sku)
        barcode_type = request.GET.get('type', 'sku')

        if barcode_type == 'ean13':
            if not product.ean13:
                return HttpResponse(
                    '<svg><text x="50%" y="50%" text-anchor="middle" fill="red">No EAN-13 available</text></svg>',
                    content_type='image/svg+xml'
                )
            # Generate EAN-13 barcode
            svg_content = generate_barcode_svg(product.ean13, format_type='ean13')
        else:
            # Generate SKU barcode (CODE128)
            svg_content = generate_barcode_svg(product.sku, format_type='code128')

        return HttpResponse(svg_content, content_type='image/svg+xml')

    except ValueError as e:
        return JsonResponse({
            "success": False,
            "error": str(e)
        }, status=400)
    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": "Error generating barcode"
        }, status=500)



@login_required
def export_categories_csv(request):
    """Export categories to CSV"""
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="categories.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Name', 'Description', 'Icon', 'Color', 'Order', 'Product Count'])
    
    categories = Category.objects.filter(is_active=True).order_by('order', 'name')
    for category in categories:
        product_count = category.products.filter(is_active=True).count()
        writer.writerow([
            category.id,
            category.name,
            category.description or '',
            category.icon,
            category.color,
            category.order,
            product_count
        ])
    
    return response


@login_required
def export_categories_excel(request):
    """Export categories to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Categories"
    
    # Headers
    headers = ['ID', 'Name', 'Description', 'Icon', 'Color', 'Order', 'Product Count']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Data
    categories = Category.objects.filter(is_active=True).order_by('order', 'name')
    for category in categories:
        product_count = category.products.filter(is_active=True).count()
        ws.append([
            category.id,
            category.name,
            category.description or '',
            category.icon,
            category.color,
            category.order,
            product_count
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="categories.xlsx"'
    wb.save(response)

    return response


@login_required
@require_http_methods(["POST"])
def import_excel(request):
    """
    Import products from Excel file
    Links categories by NAME (normalized, first letter capitalized), not by ID
    """
    if 'file' not in request.FILES:
        return JsonResponse({'error': 'No file provided'}, status=400)

    file = request.FILES['file']

    # Validate file extension
    if not (file.name.endswith('.xlsx') or file.name.endswith('.xls')):
        return JsonResponse({'error': 'Invalid file format. Please upload an Excel file.'}, status=400)

    try:
        from openpyxl import load_workbook

        # Load workbook
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        # Get header row
        headers = [cell.value for cell in ws[1]]

        # Create header mapping (case-insensitive)
        header_map = {h.lower(): i for i, h in enumerate(headers) if h}

        created_count = 0
        updated_count = 0
        errors = []

        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Helper function to get cell value by header name
                    def get_value(header_name, default=''):
                        idx = header_map.get(header_name.lower())
                        if idx is not None and idx < len(row):
                            val = row[idx]
                            return str(val).strip() if val is not None else default
                        return default

                    # Extract and clean data
                    name = get_value('Name')
                    sku = get_value('SKU')

                    if not name or not sku:
                        errors.append(f"Row {row_num}: Name and SKU are required")
                        continue

                    # Extract other fields
                    description = get_value('Description')
                    price = Decimal(get_value('Price', '0'))
                    cost = Decimal(get_value('Cost', '0'))
                    stock = int(float(get_value('Stock', '0')))
                    low_stock_threshold = int(float(get_value('Low Stock Threshold', '10')))
                    ean13 = get_value('EAN-13')

                    # Check if product exists (by SKU)
                    product, created = Product.objects.get_or_create(
                        sku=sku,
                        defaults={
                            'name': name,
                            'description': description,
                            'price': price,
                            'cost': cost,
                            'stock': stock,
                            'low_stock_threshold': low_stock_threshold,
                            'ean13': ean13 if ean13 else '',
                        }
                    )

                    if not created:
                        # Update existing product
                        product.name = name
                        product.description = description
                        product.price = price
                        product.cost = cost
                        product.stock = stock
                        product.low_stock_threshold = low_stock_threshold
                        if ean13:
                            product.ean13 = ean13
                        product.save()
                        updated_count += 1
                    else:
                        created_count += 1

                    # Handle categories - LINK BY NAME (normalized)
                    categories_str = get_value('Categories')
                    if categories_str:
                        # Clear existing categories
                        product.categories.clear()

                        # Split by comma and process each category
                        category_names = [c.strip() for c in categories_str.split(',') if c.strip()]

                        for cat_name in category_names:
                            # Normalize: first letter capitalized, rest lowercase
                            normalized_name = cat_name.capitalize()

                            # Find category (case-insensitive lookup)
                            category = Category.objects.filter(name__iexact=normalized_name).first()
                            if not category:
                                # Create if doesn't exist
                                category = Category.objects.create(name=normalized_name)

                            # Add to product
                            product.categories.add(category)

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    continue

        # Prepare response message
        message = f"Import completed: {created_count} products created, {updated_count} updated"
        if errors:
            message += f". {len(errors)} errors occurred."

        return JsonResponse({
            'success': True,
            'message': message,
            'created': created_count,
            'updated': updated_count,
            'errors': errors
        })

    except Exception as e:
        return JsonResponse({'error': f'Error processing file: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def import_categories_csv(request):
    """
    Import categories from CSV file
    Links categories by NAME (normalized, first letter capitalized)
    """
    if 'file' not in request.FILES:
        return JsonResponse({'error': 'No file provided'}, status=400)

    file = request.FILES['file']

    # Validate file extension
    if not file.name.endswith('.csv'):
        return JsonResponse({'error': 'Invalid file format. Please upload a CSV file.'}, status=400)

    try:
        # Read CSV file
        decoded_file = file.read().decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(decoded_file))

        created_count = 0
        updated_count = 0
        errors = []

        with transaction.atomic():
            for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 (header is row 1)
                try:
                    # Extract and clean data
                    name = row.get('Name', '').strip()

                    if not name:
                        errors.append(f"Row {row_num}: Name is required")
                        continue

                    # Normalize name: first letter capitalized
                    normalized_name = name.capitalize()

                    # Extract other fields
                    description = row.get('Description', '').strip()
                    icon = row.get('Icon', 'pricetag-outline').strip()
                    color = row.get('Color', 'primary').strip()
                    order = int(row.get('Order', 100))

                    # Check if category exists (by normalized name, case-insensitive)
                    category = Category.objects.filter(name__iexact=normalized_name).first()

                    if category:
                        # Update existing category
                        category.description = description
                        category.icon = icon
                        category.color = color
                        category.order = order
                        category.save()
                        updated_count += 1
                    else:
                        # Create new category
                        Category.objects.create(
                            name=normalized_name,
                            description=description,
                            icon=icon,
                            color=color,
                            order=order
                        )
                        created_count += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    continue

        # Prepare response message
        message = f"Import completed: {created_count} categories created, {updated_count} updated"
        if errors:
            message += f". {len(errors)} errors occurred."

        return JsonResponse({
            'success': True,
            'message': message,
            'created': created_count,
            'updated': updated_count,
            'errors': errors
        })

    except Exception as e:
        return JsonResponse({'error': f'Error processing file: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def import_categories_excel(request):
    """
    Import categories from Excel file
    Links categories by NAME (normalized, first letter capitalized)
    """
    if 'file' not in request.FILES:
        return JsonResponse({'error': 'No file provided'}, status=400)

    file = request.FILES['file']

    # Validate file extension
    if not (file.name.endswith('.xlsx') or file.name.endswith('.xls')):
        return JsonResponse({'error': 'Invalid file format. Please upload an Excel file.'}, status=400)

    try:
        from openpyxl import load_workbook

        # Load workbook
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        # Get header row
        headers = [cell.value for cell in ws[1]]

        # Create header mapping (case-insensitive)
        header_map = {h.lower(): i for i, h in enumerate(headers) if h}

        created_count = 0
        updated_count = 0
        errors = []

        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Helper function to get cell value by header name
                    def get_value(header_name, default=''):
                        idx = header_map.get(header_name.lower())
                        if idx is not None and idx < len(row):
                            val = row[idx]
                            return str(val).strip() if val is not None else default
                        return default

                    # Extract and clean data
                    name = get_value('Name')

                    if not name:
                        errors.append(f"Row {row_num}: Name is required")
                        continue

                    # Normalize name: first letter capitalized
                    normalized_name = name.capitalize()

                    # Extract other fields
                    description = get_value('Description')
                    icon = get_value('Icon', 'pricetag-outline')
                    color = get_value('Color', 'primary')
                    order = int(float(get_value('Order', '100')))

                    # Check if category exists (by normalized name, case-insensitive)
                    category = Category.objects.filter(name__iexact=normalized_name).first()

                    if category:
                        # Update existing category
                        category.description = description
                        category.icon = icon
                        category.color = color
                        category.order = order
                        category.save()
                        updated_count += 1
                    else:
                        # Create new category
                        Category.objects.create(
                            name=normalized_name,
                            description=description,
                            icon=icon,
                            color=color,
                            order=order
                        )
                        created_count += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    continue

        # Prepare response message
        message = f"Import completed: {created_count} categories created, {updated_count} updated"
        if errors:
            message += f". {len(errors)} errors occurred."

        return JsonResponse({
            'success': True,
            'message': message,
            'created': created_count,
            'updated': updated_count,
            'errors': errors
        })

    except Exception as e:
        return JsonResponse({'error': f'Error processing file: {str(e)}'}, status=500)
